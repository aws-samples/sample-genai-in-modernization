"""
Excel Export Module - VM to EC2 Mapping
Generates detailed Excel spreadsheet with VM analysis and AWS recommendations
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from config import output_folder_dir_path

def create_vm_to_ec2_mapping_excel(detailed_results_df, output_filename='vm_to_ec2_mapping.xlsx'):
    """
    Create detailed Excel spreadsheet with VM to EC2 mapping
    
    Args:
        detailed_results_df: DataFrame from pricing calculator with VM and EC2 details
        output_filename: Output Excel filename
    
    Returns:
        Path to generated Excel file
    """
    
    # Prepare data for Excel
    excel_data = []
    
    for idx, row in detailed_results_df.iterrows():
        excel_row = {
            # On-Premises VM Details
            'VM Name': row.get('vm_name', f'VM-{idx+1}'),
            'VM vCPU (Provisioned)': row.get('original_vcpu', row.get('vcpu', 0)),
            'VM Memory GB (Provisioned)': row.get('original_memory_gb', row.get('memory_gb', 0)),
            'VM Storage GB (Provisioned)': row.get('original_storage_gb', row.get('storage_gb', 0)),
            'VM OS': row.get('os', 'Unknown'),
            
            # Utilization Data (if available)
            'CPU Utilization %': row.get('cpu_util', 'N/A'),
            'Memory Utilization %': row.get('memory_util', 'N/A'),
            'Storage Used GB': row.get('storage_used_gb', 'N/A'),
            
            # Right-Sizing Applied
            'Right-Sizing Applied': 'Yes' if row.get('right_sizing_applied', False) else 'No',
            'vCPU Reduction %': row.get('vcpu_reduction', 0) if row.get('right_sizing_applied', False) else 0,
            'Memory Reduction %': row.get('memory_reduction', 0) if row.get('right_sizing_applied', False) else 0,
            'Storage Reduction %': row.get('storage_reduction', 0) if row.get('right_sizing_applied', False) else 0,
            
            # Right-Sized Specs (After Optimization)
            'Optimized vCPU': row.get('vcpu', 0),
            'Optimized Memory GB': row.get('memory_gb', 0),
            'Optimized Storage GB': row.get('storage_gb', 0),
            
            # AWS EC2 Recommendation
            'AWS Instance Type': row.get('instance_type', 'N/A'),
            'EC2 vCPU': _get_instance_vcpu(row.get('instance_type', '')),
            'EC2 Memory GB': _get_instance_memory(row.get('instance_type', '')),
            'EC2 OS Type': row.get('os_type', 'Linux'),
            
            # AWS Pricing (3-Year No Upfront RI)
            'EC2 Hourly Rate ($)': row.get('hourly_rate', 0),
            'EC2 Monthly Cost ($)': row.get('monthly_compute', 0),
            'EBS Storage GB': row.get('storage_gb', 0),
            'EBS Monthly Cost ($)': row.get('monthly_storage', 0),
            'Data Transfer Monthly ($)': row.get('monthly_data_transfer', 0),
            'Total Monthly Cost ($)': row.get('monthly_total', 0),
            'Total Annual Cost ($)': row.get('monthly_total', 0) * 12,
        }
        
        excel_data.append(excel_row)
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Generate Excel file with formatting
    output_path = os.path.join(output_folder_dir_path, output_filename)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='VM to EC2 Mapping', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['VM to EC2 Mapping']
        
        # Apply formatting
        _format_worksheet(worksheet, len(df))
        
        # Add summary sheet
        _add_summary_sheet(workbook, df)
    
    print(f"✓ Excel export created: {output_path}")
    return output_path

def _get_instance_vcpu(instance_type):
    """Get vCPU count for instance type"""
    from aws_pricing_calculator import AWSPricingCalculator
    calculator = AWSPricingCalculator()
    specs = calculator.INSTANCE_SPECS.get(instance_type, (0, 0))
    return specs[0]

def _get_instance_memory(instance_type):
    """Get memory GB for instance type"""
    from aws_pricing_calculator import AWSPricingCalculator
    calculator = AWSPricingCalculator()
    specs = calculator.INSTANCE_SPECS.get(instance_type, (0, 0))
    return specs[1]

def _format_worksheet(worksheet, num_rows):
    """Apply formatting to worksheet"""
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    section_fills = {
        'vm': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'util': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'rightsizing': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'optimized': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'ec2': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'pricing': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    }
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths and apply section colors
    column_sections = [
        (1, 5, 'vm'),      # VM Name to VM OS
        (6, 8, 'util'),    # Utilization data
        (9, 12, 'rightsizing'),  # Right-sizing info
        (13, 15, 'optimized'),   # Optimized specs
        (16, 19, 'ec2'),   # EC2 details
        (20, 25, 'pricing'),  # Pricing details
    ]
    
    for col_start, col_end, section in column_sections:
        for col in range(col_start, col_end + 1):
            # Set column width
            worksheet.column_dimensions[worksheet.cell(1, col).column_letter].width = 18
            
            # Apply section color to data rows
            for row in range(2, num_rows + 2):
                cell = worksheet.cell(row, col)
                cell.fill = section_fills[section]
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Format numbers
                if 'Cost' in worksheet.cell(1, col).value or 'Rate' in worksheet.cell(1, col).value:
                    cell.number_format = '$#,##0.00'
                elif '%' in worksheet.cell(1, col).value:
                    cell.number_format = '0.0'
                elif 'GB' in worksheet.cell(1, col).value or 'vCPU' in worksheet.cell(1, col).value:
                    cell.number_format = '#,##0.0'
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

def _add_summary_sheet(workbook, df):
    """Add summary sheet with aggregated statistics"""
    
    summary_sheet = workbook.create_sheet('Summary', 0)
    
    # Calculate summary statistics
    total_vms = len(df)
    total_monthly_cost = df['Total Monthly Cost ($)'].sum()
    total_annual_cost = df['Total Annual Cost ($)'].sum()
    
    # Right-sizing statistics
    vms_right_sized = len(df[df['Right-Sizing Applied'] == 'Yes'])
    avg_vcpu_reduction = df[df['Right-Sizing Applied'] == 'Yes']['vCPU Reduction %'].mean() if vms_right_sized > 0 else 0
    avg_memory_reduction = df[df['Right-Sizing Applied'] == 'Yes']['Memory Reduction %'].mean() if vms_right_sized > 0 else 0
    avg_storage_reduction = df[df['Right-Sizing Applied'] == 'Yes']['Storage Reduction %'].mean() if vms_right_sized > 0 else 0
    
    # Instance type distribution
    instance_counts = df['AWS Instance Type'].value_counts()
    
    # OS distribution
    os_counts = df['EC2 OS Type'].value_counts()
    
    # Write summary data
    summary_data = [
        ['AWS Migration Cost Analysis - Summary'],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        [''],
        ['Overall Statistics'],
        ['Total VMs:', total_vms],
        ['Total Monthly Cost:', f'${total_monthly_cost:,.2f}'],
        ['Total Annual Cost (ARR):', f'${total_annual_cost:,.2f}'],
        ['Average Cost per VM (Monthly):', f'${total_monthly_cost/total_vms:,.2f}' if total_vms > 0 else '$0'],
        [''],
        ['Right-Sizing Statistics'],
        ['VMs Right-Sized:', vms_right_sized],
        ['VMs Not Right-Sized:', total_vms - vms_right_sized],
        ['Average vCPU Reduction:', f'{avg_vcpu_reduction:.1f}%'],
        ['Average Memory Reduction:', f'{avg_memory_reduction:.1f}%'],
        ['Average Storage Reduction:', f'{avg_storage_reduction:.1f}%'],
        [''],
        ['OS Distribution'],
    ]
    
    for os_type, count in os_counts.items():
        summary_data.append([f'{os_type} VMs:', count, f'({count/total_vms*100:.1f}%)'])
    
    summary_data.append([''])
    summary_data.append(['Top 10 Instance Types'])
    
    for instance_type, count in instance_counts.head(10).items():
        monthly_cost = df[df['AWS Instance Type'] == instance_type]['Total Monthly Cost ($)'].sum()
        summary_data.append([instance_type, f'{count} VMs', f'${monthly_cost:,.2f}/month'])
    
    # Write to sheet
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row_idx, col_idx, value)
            
            # Format headers
            if row_idx in [1, 4, 10, 18, 20]:
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            elif col_idx == 1 and row_idx > 1:
                cell.font = Font(bold=True)
    
    # Set column widths
    summary_sheet.column_dimensions['A'].width = 35
    summary_sheet.column_dimensions['B'].width = 20
    summary_sheet.column_dimensions['C'].width = 20

def export_vm_to_ec2_mapping(pricing_results, output_filename='vm_to_ec2_mapping.xlsx'):
    """
    Main export function - creates Excel with VM to EC2 mapping
    
    Args:
        pricing_results: Results dictionary from pricing calculator
        output_filename: Output filename
    
    Returns:
        Path to generated Excel file
    """
    detailed_df = pricing_results.get('detailed_results')
    
    if detailed_df is None or len(detailed_df) == 0:
        print("⚠ No detailed results available for Excel export")
        return None
    
    return create_vm_to_ec2_mapping_excel(detailed_df, output_filename)


if __name__ == "__main__":
    print("Excel Export Module - VM to EC2 Mapping")
    print("Use export_vm_to_ec2_mapping() to generate Excel reports")


def export_rvtools_dual_pricing(results_option1, results_option2, output_filename='vm_to_ec2_mapping.xlsx'):
    """
    Export RVTools pricing with BOTH pricing options to Excel
    
    Similar to IT Inventory dual pricing, but EC2-only (no RDS)
    - Option 1: 3-Year EC2 Instance Savings Plan
    - Option 2: 3-Year Compute Savings Plan
    
    Args:
        results_option1: Results from Option 1 (EC2 Instance SP)
        results_option2: Results from Option 2 (Compute SP)
        output_filename: Output Excel filename
    
    Returns:
        Path to generated Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd
    from datetime import datetime
    
    output_path = os.path.join(output_folder_dir_path, output_filename)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Tab 1: Pricing Comparison Summary
    monthly_savings = results_option2['summary']['total_monthly_cost'] - results_option1['summary']['total_monthly_cost']
    annual_savings = monthly_savings * 12
    three_year_savings = monthly_savings * 36
    savings_pct = (monthly_savings / results_option2['summary']['total_monthly_cost'] * 100) if results_option2['summary']['total_monthly_cost'] > 0 else 0
    
    comparison_data = {
        'Metric': [
            'Total VMs',
            '',
            'Option 1: 3-Year EC2 Instance Savings Plan',
            'Total Monthly Cost',
            'Total Annual Cost',
            '3-Year Pricing',
            '',
            'Option 2: 3-Year Compute Savings Plan',
            'Total Monthly Cost',
            'Total Annual Cost',
            '3-Year Pricing',
            '',
            'Savings (Option 1 vs Option 2)',
            'Monthly Savings',
            'Annual Savings',
            '3-Year Savings',
            'Savings Percentage',
            '',
            'Region',
            'Recommendation'
        ],
        'Value': [
            results_option1['summary']['total_vms'],
            '',
            '',
            f"${results_option1['summary']['total_monthly_cost']:,.2f}",
            f"${results_option1['summary']['total_arr']:,.2f}",
            f"${results_option1['summary']['total_monthly_cost'] * 36:,.2f}",
            '',
            '',
            f"${results_option2['summary']['total_monthly_cost']:,.2f}",
            f"${results_option2['summary']['total_arr']:,.2f}",
            f"${results_option2['summary']['total_monthly_cost'] * 36:,.2f}",
            '',
            '',
            f"${monthly_savings:,.2f}",
            f"${annual_savings:,.2f}",
            f"${three_year_savings:,.2f}",
            f"{savings_pct:.2f}%",
            '',
            results_option1['summary']['region'],
            f"Option 1 saves ${monthly_savings:,.2f}/month ({savings_pct:.1f}%)"
        ]
    }
    
    df_comparison = pd.DataFrame(comparison_data)
    ws_comparison = wb.create_sheet('Pricing Comparison')
    
    # Write header row
    ws_comparison.cell(1, 1, 'Metric')
    ws_comparison.cell(1, 2, 'Value')
    ws_comparison['A1'].font = Font(bold=True, color='FFFFFF')
    ws_comparison['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws_comparison['B1'].font = Font(bold=True, color='FFFFFF')
    ws_comparison['B1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data rows (starting from row 2)
    for r_idx, row in enumerate(df_comparison.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_comparison.cell(r_idx, c_idx, value)
            
            # Format section headers (adjusted for header row)
            # Data indices: 2='Option 1', 7='Option 2', 12='Savings'
            # With header row: 2+2=4, 7+2=9, 12+2=14
            if r_idx in [4, 9, 14]:  # Section header rows
                cell.font = Font(bold=True, size=11, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            elif c_idx == 1 and r_idx > 1:  # Don't bold the header row again
                cell.font = Font(bold=True)
    
    ws_comparison.column_dimensions['A'].width = 40
    ws_comparison.column_dimensions['B'].width = 25
    
    # Tab 2: EC2 Details (Option 1 - EC2 Instance SP)
    ec2_details_option1 = []
    detailed_df_option1 = results_option1['detailed_results']
    
    for idx, row in detailed_df_option1.iterrows():
        from aws_pricing_calculator import AWSPricingCalculator
        calculator = AWSPricingCalculator()
        instance_specs = calculator.INSTANCE_SPECS.get(row['instance_type'], (0, 0))
        
        ec2_details_option1.append({
            'VM Name': row['vm_name'],
            'VM vCPU': row.get('original_vcpu', row['vcpu']),
            'VM Memory GB': row.get('original_memory_gb', row['memory_gb']),
            'VM Storage GB': row.get('original_storage_gb', row['storage_gb']),
            'VM OS': row['os'],
            'Right-Sizing Applied': 'Yes' if row.get('right_sizing_applied', False) else 'No',
            'Optimized vCPU': row['vcpu'],
            'Optimized Memory GB': row['memory_gb'],
            'Optimized Storage GB': row['storage_gb'],
            'AWS Instance Type': row['instance_type'],
            'EC2 vCPU': instance_specs[0],
            'EC2 Memory GB': instance_specs[1],
            'EC2 OS Type': row['os_type'],
            'Pricing Model': '3-Year EC2 Instance SP',
            'EC2 Hourly Rate ($)': row['hourly_rate'],
            'EC2 Monthly Compute ($)': row['monthly_compute'],
            'EBS Monthly Storage ($)': row['monthly_storage'],
            'Data Transfer Monthly ($)': row['monthly_data_transfer'],
            'Total Monthly Cost ($)': row['monthly_total'],
            'Total Annual Cost ($)': row['monthly_total'] * 12
        })
    
    df_ec2_option1 = pd.DataFrame(ec2_details_option1)
    ws_ec2_option1 = wb.create_sheet('EC2 Details - Option 1')
    
    # Write headers first
    for c_idx, col_name in enumerate(df_ec2_option1.columns, 1):
        cell = ws_ec2_option1.cell(1, c_idx, col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data starting from row 2
    for r_idx, row in enumerate(df_ec2_option1.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_ec2_option1.cell(r_idx, c_idx, value)
    
    # Tab 3: EC2 Details (Option 2 - Compute SP)
    ec2_details_option2 = []
    detailed_df_option2 = results_option2['detailed_results']
    
    for idx, row in detailed_df_option2.iterrows():
        from aws_pricing_calculator import AWSPricingCalculator
        calculator = AWSPricingCalculator()
        instance_specs = calculator.INSTANCE_SPECS.get(row['instance_type'], (0, 0))
        
        ec2_details_option2.append({
            'VM Name': row['vm_name'],
            'VM vCPU': row.get('original_vcpu', row['vcpu']),
            'VM Memory GB': row.get('original_memory_gb', row['memory_gb']),
            'VM Storage GB': row.get('original_storage_gb', row['storage_gb']),
            'VM OS': row['os'],
            'Right-Sizing Applied': 'Yes' if row.get('right_sizing_applied', False) else 'No',
            'Optimized vCPU': row['vcpu'],
            'Optimized Memory GB': row['memory_gb'],
            'Optimized Storage GB': row['storage_gb'],
            'AWS Instance Type': row['instance_type'],
            'EC2 vCPU': instance_specs[0],
            'EC2 Memory GB': instance_specs[1],
            'EC2 OS Type': row['os_type'],
            'Pricing Model': '3-Year Compute SP',
            'EC2 Hourly Rate ($)': row['hourly_rate'],
            'EC2 Monthly Compute ($)': row['monthly_compute'],
            'EBS Monthly Storage ($)': row['monthly_storage'],
            'Data Transfer Monthly ($)': row['monthly_data_transfer'],
            'Total Monthly Cost ($)': row['monthly_total'],
            'Total Annual Cost ($)': row['monthly_total'] * 12
        })
    
    df_ec2_option2 = pd.DataFrame(ec2_details_option2)
    ws_ec2_option2 = wb.create_sheet('EC2 Details - Option 2')
    
    # Write headers first
    for c_idx, col_name in enumerate(df_ec2_option2.columns, 1):
        cell = ws_ec2_option2.cell(1, c_idx, col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data starting from row 2
    for r_idx, row in enumerate(df_ec2_option2.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_ec2_option2.cell(r_idx, c_idx, value)
    
    # Tab 4: EC2 Comparison (Option 1 vs Option 2)
    ec2_comparison = []
    for idx, row_option1 in detailed_df_option1.iterrows():
        row_option2 = detailed_df_option2.iloc[idx]
        
        savings = row_option2['monthly_total'] - row_option1['monthly_total']
        savings_pct = (savings / row_option2['monthly_total'] * 100) if row_option2['monthly_total'] > 0 else 0
        
        ec2_comparison.append({
            'VM Name': row_option1['vm_name'],
            'Instance Type': row_option1['instance_type'],
            'OS Type': row_option1['os_type'],
            'Option 1 Monthly ($)': row_option1['monthly_total'],
            'Option 2 Monthly ($)': row_option2['monthly_total'],
            'Monthly Savings ($)': savings,
            'Savings %': f"{savings_pct:.2f}%"
        })
    
    df_ec2_comparison = pd.DataFrame(ec2_comparison)
    ws_ec2_comparison = wb.create_sheet('EC2 Comparison')
    
    # Write headers first
    for c_idx, col_name in enumerate(df_ec2_comparison.columns, 1):
        cell = ws_ec2_comparison.cell(1, c_idx, col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data starting from row 2
    for r_idx, row in enumerate(df_ec2_comparison.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_ec2_comparison.cell(r_idx, c_idx, value)
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel export created: {output_path}")
    
    return output_path
